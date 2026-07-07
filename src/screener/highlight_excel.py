from openpyxl import load_workbook
from openpyxl.styles import PatternFill
wb = load_workbook("output/screener_output.xlsx")
ws = wb.active
green = PatternFill(start_color="90EE90",end_color="90EE90",fill_type="solid")
for row in range(2, 12):
    for cell in ws[row]:
        cell.fill = green
wb.save("output/screener_output.xlsx")
print("Top 10 companies highlighted.")