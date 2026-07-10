from openpyxl import load_workbook
from openpyxl.styles import PatternFill
wb = load_workbook("output/peer_comparison.xlsx")
green = PatternFill(start_color="90EE90",end_color="90EE90",fill_type="solid")
gold = PatternFill(start_color="FFD700",end_color="FFD700",fill_type="solid")
for ws in wb.worksheets:
    for row in range(2, 7):
        for cell in ws[row]:
            cell.fill = green
    last = ws.max_row
    for cell in ws[last]:
        cell.fill = gold
wb.save("output/peer_comparison.xlsx")
print("Formatting Completed.")